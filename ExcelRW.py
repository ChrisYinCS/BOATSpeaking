# -*- coding: utf-8 -*-
import openpyxl
import os
import time


MAX_SCORE_INDAY = 10    # the max score points that a member can get is 10
SCORE_FILE_NAME = 'Score.xlsm'  # xlsm format keeps the VBA code
SCORE_FILE_PATH = 'E:\\PythonProjects\\BOATSpeaking\\'
PATH_AND_NAME = SCORE_FILE_PATH + SCORE_FILE_NAME

def excel_init(members_dict):
    """initialize the excel spreadsheet

    :param members_dict: dict of members wit their names and group number

    """
    if os.path.exists(PATH_AND_NAME):
        openpyxl.load_workbook(PATH_AND_NAME)
    else:
        wb = openpyxl.Workbook()
        write_basic_info(wb, member_dict=members_dict)
        wb.save(PATH_AND_NAME)  # the first time to save, need path


def excel_save(index, score):
    """save score in excel

    :param index: the member index
    :param score: the score of that member

    """
    if os.path.exists(PATH_AND_NAME):
        wb = openpyxl.load_workbook(PATH_AND_NAME)
        wb.guess_types = True
        ws = wb.active
        row1 = ws[1]
        temp = len(row1)    # it will return the number of written cells in row1
        # if it is the first one to finish task in a day, need to write the current day in excel
        if ws.cell(row=1, column=temp).value !=  time.strftime("%Y-%m-%d",time.localtime()):
            ws.cell(row=1, column=temp+1).value = time.strftime("%Y-%m-%d",time.localtime())
            ws.cell(row=1, column=temp+1).number_format = 'yy-mm-dd'
            write_score(worksheet=ws, row_number= index+2, column_number=temp+1, score=score)
        else:   # if it is not the first one to finish the task in a day
            write_score(worksheet=ws, row_number=index+2, column_number=temp, score=score)

        wb.save(SCORE_FILE_NAME)


def write_score(worksheet, row_number:'int > 0', column_number:'int > 0', score: 'int >= 0'):
    """write score in a cell

     :param worksheet: the activated worksheet
     :param row_number: row number of the cell  type: int
     :param column_number: column number of that cell   type:int
     :param score: score to be written in the cell

     """
    # if this member has already finished task today and he/she is doing again
    if worksheet.cell(row_number, column_number).value is None:
        temp = score
        if temp > MAX_SCORE_INDAY:
            temp = MAX_SCORE_INDAY
    else:
        temp = worksheet.cell(row_number, column_number).value + score
        if temp > MAX_SCORE_INDAY:
            temp = MAX_SCORE_INDAY

    worksheet.cell(row_number, column_number).value = temp
    worksheet.cell(row_number, column_number).number_format = '0;[Red]0'  # format is number with no digits


def read_score(member_index:'int', total_members):
    """write score in a cell

     :param member_index: index of that member
     :param total_members: total number of members

     """
    if os.path.exists(PATH_AND_NAME):
        score_values = []

        # must set "data_only" as True, otherwise the formula itself will be returned but not its' value
        wb = openpyxl.load_workbook(PATH_AND_NAME, data_only=True)
        ws = wb.active
        row1 = ws[1]
        temp = len(row1)
        member_score = ws.cell(row=member_index+2, column=temp).value
        score_values.append(member_score)
        group_number = ws.cell(row=member_index+2, column=2).value
        # !!!!!Important note: the bolow code will always return a "NoneType" if
        # the excel file is not MANUALLY opened and saved before!!!!!!! This is because of openpyxl shortage
        # see web: https://stackoverflow.com/questions/35681902/openpyxl-load-workbookfile-data-only-true-doenst-work
        group_score = ws.cell(row=group_number+total_members+3, column=4).value
        score_values.append(group_score)

        return score_values


def write_basic_info(wb, member_dict):
    """write score in a cell

     :param wb: the activated worksheet
     :param member_dict: dict of members with their names and group number

     """
    ws = wb.active
    total_member = len(member_dict)
    name_list = list(member_dict)
    group_list = list(member_dict.values())
    group_set = list(set(group_list)) # how many different groups there are

    ws['A1'] = 'Participant Name'
    ws['B1'] = 'Group Number'
    ws['C1'] = 'Starting score'
    ws['D1'] = 'Total score'

    ws['A%d'%(total_member+3)] = 'Group Number'
    ws['B%d'%(total_member+3)] = 'Starting Score'
    ws['C%d'%(total_member+3)] = 'Total Group Score'

    for i in range(2, total_member+2):
        ws.cell(row=i, column=1).value = name_list[i-2]
        ws.cell(row=i, column=2).value = group_list[i-2]
        ws.cell(row=i, column=2).number_format = '0;[Red]0'
        ws.cell(row=i, column=3).value = 0      # starting score is 0 in default
        ws.cell(row=i, column=3).number_format = '0;[Red]0'  # starting score is 0 in default

# calculated cells
    for i in range(2, total_member+2):    # total score of each member
        ws.cell(row=i, column=4).value = '=SUM(E%d:XFD%d)'%(i,i)
        ws.cell(row=i, column=4).number_format = '0;[Red]0'

    for i in range(len(group_set)):     # set of group
        ws.cell(row=total_member + 4 + i, column=1).value = group_set[i]
        ws.cell(row=total_member + 4 + i, column=1).number_format = '0;[Red]0'


        # find the members that belongs to this group, calculate their total starting score
        ws.cell(row=total_member + 4 + i, column=2).value = \
            '=SUMIF(B2:B%d, A%d, C2:C%d)'%(total_member+1,total_member+4+i,total_member+1)
        ws.cell(row=total_member + 4 + i, column=2).number_format = '0;[Red]0'

        #  find the members that belongs to this group, calculate their total score
        ws.cell(row=total_member + 4 + i, column=3).value = \
            '=SUM(B%d,SUMIF(B2:B%d, A%d, D2:D%d))'%(total_member+4+i,total_member+1,total_member+4+i,total_member+1)
        ws.cell(row=total_member + 4 + i, column=3).number_format = '0;[Red]0'